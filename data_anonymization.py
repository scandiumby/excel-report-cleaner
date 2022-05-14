import random

from openpyxl import load_workbook
from russian_names import RussianNames

DISCOVERY_COL = 2

if __name__ == "__main__":
    workbook = load_workbook(filename='data/input.xlsx')

    sheet = workbook.active
    processed_cells = list(sheet.rows)[8:]
    names = RussianNames(count=500, rare=True, patronymic_reduction=True).get_batch()
    name_index = 0

    for row in processed_cells:
        current_row_value = row[DISCOVERY_COL].value
        print(f"{current_row_value=}")
        if current_row_value is not None:
            sheet.cell(row=row[0].row, column=DISCOVERY_COL+1, value=names[name_index])
            sheet.cell(row=row[0].row, column=DISCOVERY_COL, value=f"Отдел #{random.randint(0, 10)}")
            name_index += 1

    workbook.save(filename='data/input.xlsx')