from openpyxl import load_workbook

DISCOVERY_COL = 2
START_ROW_INDEX = 8


def clear_data(input_filename: str = "data/input.xlsx", output_filename: str = "data/output.xlsx") -> None:
    workbook = load_workbook(filename=input_filename)

    sheet = workbook.active
    processed_cells = list(sheet.rows)[START_ROW_INDEX:]
    current_row_index = processed_cells[0][0].row

    for row in processed_cells:
        current_row_value = row[DISCOVERY_COL].value
        next_row_value = sheet.cell(row=current_row_index + 1, column=DISCOVERY_COL).value
        if current_row_value is None and next_row_value is None:
            sheet.delete_rows(current_row_index)
            current_row_index -= 1
        current_row_index += 1

    workbook.save(filename=output_filename)


if __name__ == '__main__':
    clear_data()
